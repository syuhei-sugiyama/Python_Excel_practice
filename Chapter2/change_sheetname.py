from openpyxl import load_workbook

# load_workbook関数でブック読み込み
# 引数にread_onlyを指定しないことで、ブック変更可能になる
wb = load_workbook('集計.xlsx')

""" 
取得した、Workbookオブジェクトからworksheets属性によって
Worksheetオブジェクトを取得して、変数wsに代入
 """
for i, ws in enumerate(wb.worksheets):
    # 取得したWorksheetオブジェクトのタイトル名に「ID_」を先頭に付与して
    # Worksheetオブジェクトのタイトル名に代入
    ws.title = 'ID_' + ws.title
    # 変数iを5で割った余りが0だった場合
    if (i + 1) % 5 == 0:
        # for文の中で、このブロックの中に入ったときに、取得していた、
        # Worksheetオブジェクトの、シートの見出しの色を変更する
        ws.sheet_properties.tabColor = '0000FF'

wb.save('集計_変更後.xlsx')
