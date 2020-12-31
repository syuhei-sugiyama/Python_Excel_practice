from pathlib import Path

from openpyxl import load_workbook, Workbook

# Workbookオブジェクト生成
wb_new = Workbook()
# 生成したWorkbookオブジェクトからWorksheetオブジェクト取得
ws_new = wb_new.active
# 取得したWorksheetオブジェクトのシート名を変更
ws_new.title = '一覧表'

# 新規ブックのセルB2に「部署名」を設定
ws_new['B2'] = '部署名'
# 新規ブックのセルC2に「氏名」を設定
ws_new['C2'] = '氏名'

# Pathオブジェクトの生成
# 引数に指定したパスのPathオブジェクトが生成される
# 今回は、「カレントレコードから、booksフォルダ配下」を指定
path = Path('./books')

# Path.globメソッドを使用すると、引数に一致したファイルのみを順に取得する
for i, file in enumerate(path.glob('*.xlsx')):
    # load_workbook関数によって、globメソッドで取得したブックを読み込む
    # 引数に、読み込むファイルと、read_only=Trueつまり読み取り専用の指定をしている
    wb = load_workbook(file, read_only=True)
    # 読み取り専用で取得したブックから、「チェックリスト」というシートを指定して、Worksheetオブジェクト取得
    # つまり、操作するシートに「チェックリスト」というシートを指定したということ
    ws = wb['チェックリスト']

    row_no = i + 3
    """ ws_newつまり新規ブックのB*セルに、読み取り専用で取得したブックの、チェックリストシートの、
    C2セルの値を設定。指定のセルから値を取得するときは、「.value」が必要 """
    ws_new[f'B{row_no}'] = ws['C2'].value
    """ C*セルも同様 """
    ws_new[f'C{row_no}'] = ws['C3'].value

# 最後に、新規ブックを、指定した名前で保存
wb_new.save('一覧表.xlsx')
