from pathlib import Path

from openpyxl import load_workbook, Workbook

# 新規ブックのWorkbookオブジェクトを取得
wb_new = Workbook()
# Workbookオブジェクトからデフォルトで作成されたシートを取得
ws_new = wb_new.active
# 取得したシート名を「集計」に変更
ws_new.title = '集計'

# 取得したWorksheetオブジェクト(シート名「集計」のシート)に対して
# 引数に指定したセルに対して、各値を設定
ws_new['B2'] = 'ブック名'
ws_new['C2'] = '全シート数'
ws_new['D2'] = '非表示シート数'

"""
Pathオブジェクトを取得
引数には、カレントディレクトリからbooksフォルダへ移動した後のパスを取得
 """
path = Path('./books')
# globメソッドで、引数に指定した、「maで始まる、拡張子がxlsx」のファイル数分だけ繰り返し
for i, file in enumerate(path.glob('ma*.xlsx')):
    # 取得したファイルのオブジェクトを、読み書き可能で取得
    wb = load_workbook(file)

    row_no = i + 3
    # 変数fileつまり、maで始まるファイルのオブジェクトから、name属性つまり、ブック名を取得
    # 新規ブックのオブジェクトのB*セルに設定
    ws_new[f'B{row_no}'] = file.name
    # ファイルオブジェクトのsheetnames属性により、ブックのシート名をリストで返す
    # len関数を用いることで、sheetnames属性で取得したリストの長さを取得出来る
    # 新規ブックのオブジェクトのC*セルに設定
    ws_new[f'C{row_no}'] = len(wb.sheetnames)

    """
    取得したファイルのオブジェクトからworksheets属性でシートを取得
    sheet_state属性によって、取得したシートが表示されているか、非表示かを取得
    SHEETSTATE_VISIBLE定数が、表示されていることを表している
    これと一致しないことで、非表示のシートを取得
    """
    hidden_worksheets = [
        ws for ws in wb.worksheets if ws.sheet_state != ws.SHEETSTATE_VISIBLE]
    # 取得した非表示のシート名のリストに対して、len関数を用いて、非表示のシート数を取得
    # D*セルに設定
    ws_new[f'D{row_no}'] = len(hidden_worksheets)

wb_new.save('シート数集計.xlsx')
